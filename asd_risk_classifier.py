#!/usr/bin/env python3
"""
ASD Risk Matrix Classifier

This script uses ChatGPT (OpenAI API) to analyze cyber events and recommend
ASD severity categories (C1-C6) and stakeholder levels based on the Australian
Signals Directorate risk matrix framework.

Usage:
    python asd_risk_classifier.py [--limit 5] [--force-reclassify] [--output-dir .]
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import sqlite3
import sys
import time
import uuid
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple, Literal

import pandas as pd
from dotenv import load_dotenv
from openai import OpenAI
from pydantic import BaseModel, Field, field_validator
from tqdm import tqdm

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler('asd_risk_classifier.log', encoding='utf-8')
    ]
)
logger = logging.getLogger(__name__)

# Valid stakeholder categories (from ASD risk matrix)
VALID_STAKEHOLDER_CATEGORIES = [
    "Member(s) of the public",
    "Small organisation(s)",
    "Sole traders",
    "Medium-sized organisation(s)",
    "Schools",
    "Local government",
    "State government",
    "Academia/R&D",
    "Large organisation(s)",
    "Supply chain",
    "Federal government",
    "Government shared services",
    "Regulated critical infrastructure",
    "National security",
    "Systems of National Significance"
]

# Grouped stakeholder categories for risk matrix display
STAKEHOLDER_GROUPS = {
    "Member(s) of the public": ["Member(s) of the public"],
    "Small organisation(s) / Sole traders": [
        "Small organisation(s)",
        "Sole traders"
    ],
    "Medium-sized organisation(s) / Schools / Local government": [
        "Medium-sized organisation(s)",
        "Schools",
        "Local government"
    ],
    "State government / Academia/R&D / Large organisation(s) / Supply chain": [
        "State government",
        "Academia/R&D",
        "Large organisation(s)",
        "Supply chain"
    ],
    "Federal government / Government shared services / Regulated critical infrastructure": [
        "Federal government",
        "Government shared services",
        "Regulated critical infrastructure"
    ],
    "National security / Systems of National Significance": [
        "National security",
        "Systems of National Significance"
    ]
}

# Valid impact types (from ASD risk matrix)
VALID_IMPACT_TYPES = [
    "Sustained disruption of essential systems and associated services",
    "Extensive compromise",
    "Isolated compromise",
    "Coordinated low-level malicious attack",
    "Low-level malicious attack",
    "Unsuccessful low-level malicious attack"
]

# Type aliases for strict validation
SeverityCategory = Literal["C1", "C2", "C3", "C4", "C5", "C6"]
StakeholderCategory = Literal[
    "Member(s) of the public",
    "Small organisation(s)",
    "Sole traders",
    "Medium-sized organisation(s)",
    "Schools",
    "Local government",
    "State government",
    "Academia/R&D",
    "Large organisation(s)",
    "Supply chain",
    "Federal government",
    "Government shared services",
    "Regulated critical infrastructure",
    "National security",
    "Systems of National Significance"
]
ImpactType = Literal[
    "Sustained disruption of essential systems and associated services",
    "Extensive compromise",
    "Isolated compromise",
    "Coordinated low-level malicious attack",
    "Low-level malicious attack",
    "Unsuccessful low-level malicious attack"
]


class ClassificationReasoning(BaseModel):
    """Reasoning for ASD risk classification."""
    severity_reasoning: str = Field(..., min_length=10)
    stakeholder_reasoning: str = Field(..., min_length=10)
    impact_reasoning: str = Field(..., min_length=10)
    information_quality: str = Field(..., min_length=10)


class ASDRiskClassification(BaseModel):
    """
    Structured output model for ASD risk classification.
    Uses strict typing to prevent LLM from returning invalid values.
    """
    severity_category: SeverityCategory = Field(
        ...,
        description="ASD severity category from C1 (most severe) to C6 (least severe)"
    )
    primary_stakeholder_category: StakeholderCategory = Field(
        ...,
        description="Primary stakeholder category affected by this incident"
    )
    impact_type: ImpactType = Field(
        ...,
        description="Type of impact according to ASD risk matrix"
    )
    reasoning: ClassificationReasoning = Field(
        ...,
        description="Detailed reasoning for each classification decision"
    )
    confidence: float = Field(
        ...,
        ge=0.0,
        le=1.0,
        description="Confidence score between 0 and 1"
    )

    @field_validator('confidence')
    @classmethod
    def validate_confidence(cls, v: float) -> float:
        """Ensure confidence is between 0 and 1."""
        if not 0.0 <= v <= 1.0:
            raise ValueError(f"Confidence must be between 0 and 1, got {v}")
        return v


class ASDRiskClassifier:
    """Classify cyber events using ASD risk matrix framework."""
    
    def __init__(self, db_path: str, model: str = "gpt-4o", api_key: Optional[str] = None):
        """Initialize the classifier."""
        self.db_path = Path(db_path)
        if not self.db_path.exists():
            raise FileNotFoundError(f"Database not found at {self.db_path}")
        
        self.model = model
        self.api_key = api_key or os.getenv("OPENAI_API_KEY")
        if not self.api_key:
            raise ValueError("OPENAI_API_KEY not found in environment variables or .env file")
        
        self.client = OpenAI(api_key=self.api_key)
        self.conn = sqlite3.connect(str(self.db_path))
        self.conn.row_factory = sqlite3.Row
        
        # Track API usage
        self.total_tokens = 0
        self.api_calls = 0
        self.cache_hits = 0
    
    def close(self):
        """Close database connection."""
        if self.conn:
            self.conn.close()
    
    def get_events(self, limit: int = 5, prioritize_unclassified: bool = True) -> List[Dict[str, Any]]:
        """
        Get active events from DeduplicatedEvents table.

        Args:
            limit: Maximum number of events to retrieve
            prioritize_unclassified: If True, prioritize events without classifications

        Returns:
            List of event dictionaries
        """
        cursor = self.conn.cursor()

        if prioritize_unclassified:
            # Get unclassified events first, then classified ones
            # This ensures we process new events before hitting the limit
            query = """
                SELECT
                    de.deduplicated_event_id,
                    de.title,
                    de.description,
                    de.summary,
                    de.event_type,
                    de.severity,
                    de.event_date,
                    de.records_affected,
                    de.victim_organization_name,
                    de.victim_organization_industry,
                    de.attacking_entity_name,
                    de.attack_method,
                    de.is_australian_event,
                    de.australian_relevance_score,
                    de.created_at,
                    CASE WHEN arc.deduplicated_event_id IS NULL THEN 0 ELSE 1 END as has_classification
                FROM DeduplicatedEvents de
                LEFT JOIN ASDRiskClassifications arc
                    ON de.deduplicated_event_id = arc.deduplicated_event_id
                WHERE de.status = 'Active'
                ORDER BY has_classification ASC, de.event_date DESC, de.created_at DESC
                LIMIT ?
            """
        else:
            # Original query - just get events by date
            query = """
                SELECT
                    deduplicated_event_id,
                    title,
                    description,
                    summary,
                    event_type,
                    severity,
                    event_date,
                    records_affected,
                    victim_organization_name,
                    victim_organization_industry,
                    attacking_entity_name,
                    attack_method,
                    is_australian_event,
                    australian_relevance_score,
                    created_at
                FROM DeduplicatedEvents
                WHERE status = 'Active'
                ORDER BY event_date DESC, created_at DESC
                LIMIT ?
            """

        cursor.execute(query, (limit,))
        events = [dict(row) for row in cursor.fetchall()]

        # Remove the has_classification field if it exists
        for event in events:
            event.pop('has_classification', None)

        logger.debug(f"Retrieved {len(events)} active events")
        return events
    
    def get_cached_classification(self, deduplicated_event_id: str) -> Optional[Dict[str, Any]]:
        """Check if classification already exists in cache."""
        cursor = self.conn.cursor()
        
        query = """
            SELECT * FROM ASDRiskClassifications
            WHERE deduplicated_event_id = ?
        """
        
        cursor.execute(query, (deduplicated_event_id,))
        row = cursor.fetchone()
        
        if row:
            # Convert database row to classification format
            cached = dict(row)
            # Parse reasoning_json to reasoning dict
            if 'reasoning_json' in cached and cached['reasoning_json']:
                try:
                    cached['reasoning'] = json.loads(cached['reasoning_json'])
                except json.JSONDecodeError:
                    logger.warning(f"Failed to parse reasoning_json for event {deduplicated_event_id}")
                    return None
            # Map confidence_score to confidence
            if 'confidence_score' in cached:
                cached['confidence'] = cached['confidence_score']
            return cached
        return None
    
    def build_prompt(self, event: Dict[str, Any]) -> str:
        """Build the prompt for ChatGPT classification."""
        
        # Format records_affected for display
        records_affected = event.get('records_affected')
        records_text = f"{records_affected:,}" if records_affected else "Unknown"
        
        # Build event context
        event_context = f"""
Event Details:
- Title: {event.get('title', 'N/A')}
- Event Type: {event.get('event_type', 'N/A')}
- Current Severity: {event.get('severity', 'N/A')}
- Event Date: {event.get('event_date', 'N/A')}
- Records Affected: {records_text}
- Victim Organization: {event.get('victim_organization_name', 'N/A')}
- Victim Industry: {event.get('victim_organization_industry', 'N/A')}
- Attacking Entity: {event.get('attacking_entity_name', 'N/A')}
- Attack Method: {event.get('attack_method', 'N/A')}
- Australian Event: {event.get('is_australian_event', 'N/A')}
- Australian Relevance Score: {event.get('australian_relevance_score', 'N/A')}

Description: {event.get('description', 'N/A')}

Summary: {event.get('summary', 'N/A')}
"""
        
        prompt = f"""You are an expert cybersecurity analyst classifying cyber incidents according to the Australian Signals Directorate (ASD) risk matrix framework.

Analyze the following cyber event and provide a classification:

{event_context}

ASD Severity Categories (C1 = most severe, C6 = least severe):
- C1: Most severe - Sustained disruption of critical national infrastructure, extensive data breaches affecting millions, national security implications
- C2: Very severe - Major disruption to essential services, large-scale data breaches (hundreds of thousands to millions), significant financial impact
- C3: Severe - Substantial impact on organizations, moderate data breaches (tens of thousands to hundreds of thousands), notable service disruption
- C4: Moderate - Limited impact on organizations, smaller data breaches (thousands to tens of thousands), minor service disruption
- C5: Low - Minimal impact, small data breaches (hundreds to thousands), isolated incidents
- C6: Least severe - Unsuccessful attacks, very small or no data breaches, minimal to no impact

IMPORTANT: Use the "Records Affected" field to inform severity:
- >1,000,000 records → Consider C1-C2
- 100,000-1,000,000 records → Consider C2-C3
- 10,000-100,000 records → Consider C3-C4
- 1,000-10,000 records → Consider C4-C5
- <1,000 records → Consider C5-C6
- Unknown/No data → Default to C6 (least severe)

Stakeholder Categories (select ONE primary category):
{chr(10).join(f"- {cat}" for cat in VALID_STAKEHOLDER_CATEGORIES)}

Impact Types:
{chr(10).join(f"- {impact}" for impact in VALID_IMPACT_TYPES)}

Instructions:
1. Analyze the event details, especially the number of records affected
2. Assign a severity category (C1-C6) based on impact and records affected
3. Select ONE primary stakeholder category that best represents the victim
4. Classify the impact type
5. Provide detailed reasoning for each classification
6. If there is insufficient information, default to C6 (least severe) and explain why

Return your response as a JSON object with this exact structure:
{{
    "severity_category": "C3",
    "primary_stakeholder_category": "Medium-sized organisation(s)",
    "impact_type": "Isolated compromise",
    "reasoning": {{
        "severity_reasoning": "Detailed explanation of why this severity was chosen, including consideration of records affected",
        "stakeholder_reasoning": "Explanation of why this stakeholder category was selected",
        "impact_reasoning": "Explanation of the impact type classification",
        "information_quality": "Assessment of available information quality and any limitations"
    }},
    "confidence": 0.85
}}
"""
        
        return prompt
    
    def classify_event(self, event: Dict[str, Any], force_reclassify: bool = False) -> Optional[Dict[str, Any]]:
        """Classify a single event using ChatGPT."""
        
        event_id = event['deduplicated_event_id']
        
        # Check cache first
        if not force_reclassify:
            cached = self.get_cached_classification(event_id)
            if cached:
                self.cache_hits += 1
                logger.debug(f"Using cached classification for event {event_id}")
                return cached
        
        # Build prompt
        prompt = self.build_prompt(event)
        
        # Call OpenAI API with retry logic and structured outputs
        max_retries = 3
        for attempt in range(max_retries):
            try:
                logger.debug(f"Calling API for event {event_id} (attempt {attempt + 1})")

                # Use structured outputs with Pydantic model
                # This forces the LLM to return valid data matching our schema
                response = self.client.beta.chat.completions.parse(
                    model=self.model,
                    messages=[
                        {
                            "role": "system",
                            "content": "You are an expert cybersecurity analyst specializing in ASD risk classification. You must classify cyber incidents according to the Australian Signals Directorate risk matrix framework using ONLY the valid categories provided."
                        },
                        {"role": "user", "content": prompt}
                    ],
                    response_format=ASDRiskClassification,
                    temperature=0.3
                )

                # Parse structured response
                parsed_classification = response.choices[0].message.parsed

                if parsed_classification is None:
                    logger.warning(f"Failed to parse classification for event {event_id}, retrying...")
                    if attempt < max_retries - 1:
                        time.sleep(2 ** attempt)  # Exponential backoff
                        continue
                    else:
                        logger.error(f"Failed to get valid classification after {max_retries} attempts")
                        return None

                # Track API usage
                self.api_calls += 1
                if hasattr(response, 'usage'):
                    self.total_tokens += response.usage.total_tokens

                # Convert Pydantic model to dict for consistency with existing code
                result = {
                    'severity_category': parsed_classification.severity_category,
                    'primary_stakeholder_category': parsed_classification.primary_stakeholder_category,
                    'impact_type': parsed_classification.impact_type,
                    'reasoning': {
                        'severity_reasoning': parsed_classification.reasoning.severity_reasoning,
                        'stakeholder_reasoning': parsed_classification.reasoning.stakeholder_reasoning,
                        'impact_reasoning': parsed_classification.reasoning.impact_reasoning,
                        'information_quality': parsed_classification.reasoning.information_quality
                    },
                    'confidence': parsed_classification.confidence
                }

                logger.debug(f"Successfully classified event {event_id}")
                return result

            except Exception as e:
                logger.error(f"API error for event {event_id} (attempt {attempt + 1}): {e}")
                if attempt < max_retries - 1:
                    time.sleep(2 ** attempt)
                    continue
                else:
                    return None

        return None
    
    def validate_classification(self, classification: Dict[str, Any]) -> bool:
        """Validate the classification response."""
        required_fields = ['severity_category', 'primary_stakeholder_category', 'impact_type', 'reasoning', 'confidence']
        
        # Check required fields
        for field in required_fields:
            if field not in classification:
                logger.warning(f"Missing required field: {field}")
                return False
        
        # Validate severity category
        if classification['severity_category'] not in ['C1', 'C2', 'C3', 'C4', 'C5', 'C6']:
            logger.warning(f"Invalid severity category: {classification['severity_category']}")
            return False
        
        # Validate stakeholder category
        if classification['primary_stakeholder_category'] not in VALID_STAKEHOLDER_CATEGORIES:
            logger.warning(f"Invalid stakeholder category: {classification['primary_stakeholder_category']}")
            return False
        
        # Validate impact type
        if classification['impact_type'] not in VALID_IMPACT_TYPES:
            logger.warning(f"Invalid impact type: {classification['impact_type']}")
            return False
        
        # Validate confidence
        confidence = classification['confidence']
        if not isinstance(confidence, (int, float)) or not (0.0 <= confidence <= 1.0):
            logger.warning(f"Invalid confidence score: {confidence}")
            return False
        
        # Validate reasoning structure
        reasoning = classification['reasoning']
        if not isinstance(reasoning, dict):
            logger.warning("Reasoning must be a dictionary")
            return False
        
        required_reasoning_fields = ['severity_reasoning', 'stakeholder_reasoning', 'impact_reasoning', 'information_quality']
        for field in required_reasoning_fields:
            if field not in reasoning:
                logger.warning(f"Missing reasoning field: {field}")
                return False
        
        return True
    
    def save_classification(self, event_id: str, classification: Dict[str, Any]) -> bool:
        """Save classification to database."""
        try:
            cursor = self.conn.cursor()
            
            classification_id = str(uuid.uuid4())
            reasoning_json = json.dumps(classification['reasoning'], ensure_ascii=False)
            
            query = """
                INSERT OR REPLACE INTO ASDRiskClassifications (
                    classification_id,
                    deduplicated_event_id,
                    severity_category,
                    primary_stakeholder_category,
                    impact_type,
                    reasoning_json,
                    confidence_score,
                    model_used,
                    updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """
            
            cursor.execute(query, (
                classification_id,
                event_id,
                classification['severity_category'],
                classification['primary_stakeholder_category'],
                classification['impact_type'],
                reasoning_json,
                classification['confidence'],
                self.model,
                datetime.now().isoformat()
            ))
            
            self.conn.commit()
            logger.debug(f"Saved classification for event {event_id}")
            return True
            
        except Exception as e:
            logger.error(f"Error saving classification for event {event_id}: {e}")
            self.conn.rollback()
            return False
    
    def process_events(self, limit: int = 5, force_reclassify: bool = False) -> List[Dict[str, Any]]:
        """Process events and generate classifications."""
        # Prioritize unclassified events to ensure they get processed first
        events = self.get_events(limit, prioritize_unclassified=not force_reclassify)
        
        if not events:
            logger.warning("No events found to process")
            return []
        
        results = []
        last_was_api_call = False
        
        for event in tqdm(events, desc="Classifying events"):
            event_id = event['deduplicated_event_id']
            
            # Check if classification is already cached
            cached_classification = None if force_reclassify else self.get_cached_classification(event_id)
            needs_api_call = force_reclassify or cached_classification is None
            
            # Classify event (will use cache if available)
            classification = self.classify_event(event, force_reclassify)
            
            if classification:
                # Only save if it's a new classification (not from cache)
                if cached_classification is None:
                    # Save new classification to database
                    if not self.save_classification(event_id, classification):
                        logger.warning(f"Failed to save classification for event {event_id}")
                        continue
                
                # Combine event data with classification
                result = {
                    'event': event,
                    'classification': classification
                }
                results.append(result)
                
                # Rate limiting: wait 1 second after API calls (not cache hits)
                if needs_api_call and last_was_api_call:
                    time.sleep(1)
                
                last_was_api_call = needs_api_call
            else:
                logger.warning(f"Failed to classify event {event_id}")
                last_was_api_call = False
        
        # Log summary
        logger.info(f"\nProcessing Summary:")
        logger.info(f"  Events processed: {len(results)}")
        logger.info(f"  API calls made: {self.api_calls}")
        logger.info(f"  Cache hits: {self.cache_hits}")
        logger.info(f"  Total tokens used: {self.total_tokens}")
        
        return results
    
    def export_results(self, results: List[Dict[str, Any]], output_dir: str = "risk_matrix") -> Tuple[str, str, List[str]]:
        """Export results to CSV and JSON."""
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        csv_file = output_path / f"asd_risk_classifications_{timestamp}.csv"
        json_file = output_path / f"asd_risk_classifications_{timestamp}.json"
        
        # Prepare data for export
        export_data = []
        for result in results:
            event = result['event']
            classification = result['classification']
            
            export_item = {
                'deduplicated_event_id': event['deduplicated_event_id'],
                'title': event.get('title'),
                'event_date': event.get('event_date'),
                'event_type': event.get('event_type'),
                'victim_organization_name': event.get('victim_organization_name'),
                'victim_organization_industry': event.get('victim_organization_industry'),
                'records_affected': event.get('records_affected'),
                'severity_category': classification['severity_category'],
                'primary_stakeholder_category': classification['primary_stakeholder_category'],
                'impact_type': classification['impact_type'],
                'confidence_score': classification['confidence'],
                'model_used': self.model,
                'created_at': datetime.now().isoformat(),
                'reasoning': classification['reasoning']  # Full reasoning for JSON
            }
            export_data.append(export_item)
        
        # Export to CSV
        import csv
        with open(csv_file, 'w', newline='', encoding='utf-8') as f:
            fieldnames = [
                'deduplicated_event_id', 'title', 'event_date', 'event_type',
                'victim_organization_name', 'victim_organization_industry',
                'records_affected', 'severity_category', 'primary_stakeholder_category',
                'impact_type', 'confidence_score', 'model_used', 'created_at'
            ]
            writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
            writer.writeheader()
            for item in export_data:
                # Remove reasoning from CSV (it's in JSON only)
                csv_item = {k: v for k, v in item.items() if k != 'reasoning'}
                writer.writerow(csv_item)
        
            logger.debug(f"Exported CSV to {csv_file}")
        
        # Export to JSON (with full reasoning)
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2, ensure_ascii=False, default=str)
        
        logger.debug(f"Exported JSON to {json_file}")
        
        # Compile and export risk matrix (all years and per year)
        excel_files = self.compile_risk_matrix(output_path)
        
        return str(csv_file), str(json_file), excel_files
    
    def compile_risk_matrix(self, output_path: Path) -> List[str]:
        """Compile risk matrix from all classifications and export to Excel.
        
        Returns list of Excel file paths (overall + per year).
        """
        excel_files = []
        
        # Compile overall risk matrix (all years)
        overall_file = self._compile_risk_matrix_for_year(output_path, "risk_matrix.xlsx", None)
        if overall_file:
            excel_files.append(str(overall_file))
        
        # Compile risk matrix for each year from 2020 onwards
        current_year = datetime.now().year
        for year in range(2020, current_year + 1):
            year_file = self._compile_risk_matrix_for_year(
                output_path, 
                f"risk_matrix_{year}.xlsx", 
                year
            )
            if year_file:
                excel_files.append(str(year_file))
        
        return excel_files
    
    def _compile_risk_matrix_for_year(self, output_path: Path, filename: str, year: Optional[int]) -> Optional[Path]:
        """Compile risk matrix for a specific year (or all years if year is None)."""
        excel_file = output_path / filename
        
        try:
            # Get all classifications from database, optionally filtered by year
            cursor = self.conn.cursor()
            
            if year is not None:
                # Join with DeduplicatedEvents to filter by event_date year
                query = """
                    SELECT 
                        arc.impact_type,
                        arc.primary_stakeholder_category,
                        COUNT(*) as count
                    FROM ASDRiskClassifications arc
                    JOIN DeduplicatedEvents de ON arc.deduplicated_event_id = de.deduplicated_event_id
                    WHERE CAST(strftime('%Y', de.event_date) AS INTEGER) = ?
                    GROUP BY arc.impact_type, arc.primary_stakeholder_category
                    ORDER BY arc.impact_type, arc.primary_stakeholder_category
                """
                cursor.execute(query, (year,))
            else:
                # Get all classifications (no year filter)
                query = """
                    SELECT 
                        impact_type,
                        primary_stakeholder_category,
                        COUNT(*) as count
                    FROM ASDRiskClassifications
                    GROUP BY impact_type, primary_stakeholder_category
                    ORDER BY impact_type, primary_stakeholder_category
                """
                cursor.execute(query)
            
            rows = cursor.fetchall()
            
            if not rows:
                year_label = f" for year {year}" if year is not None else ""
                logger.debug(f"No classifications found{year_label}")
                # Create empty matrix
                self._create_empty_risk_matrix(excel_file)
                return excel_file
            
            # Create DataFrame
            df = pd.DataFrame(rows, columns=['impact_type', 'primary_stakeholder_category', 'count'])
            
            # Create pivot table: impact_type (rows) x grouped stakeholder categories (columns)
            # Show only total counts (not severity codes)
            pivot_data = []
            
            for impact_type in VALID_IMPACT_TYPES:
                row_data = {'Impact Type': impact_type}
                
                # Process each stakeholder group
                for group_name, group_categories in STAKEHOLDER_GROUPS.items():
                    # Sum counts for all categories in this group
                    total_count = 0
                    for stakeholder in group_categories:
                        subset = df[
                            (df['impact_type'] == impact_type) & 
                            (df['primary_stakeholder_category'] == stakeholder)
                        ]
                        if len(subset) > 0:
                            total_count += int(subset.iloc[0]['count'])
                    
                    # Store count (only if > 0)
                    row_data[group_name] = total_count if total_count > 0 else ''
                
                pivot_data.append(row_data)
            
            # Create DataFrame for risk matrix
            risk_matrix_df = pd.DataFrame(pivot_data)
            
            # Reorder columns: Impact Type first, then grouped stakeholder categories
            columns = ['Impact Type'] + list(STAKEHOLDER_GROUPS.keys())
            risk_matrix_df = risk_matrix_df[columns]
            
            # Write to Excel with formatting
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                risk_matrix_df.to_excel(writer, sheet_name='Risk Matrix', index=False)
                
                # Get workbook and worksheet for formatting
                workbook = writer.book
                worksheet = writer.sheets['Risk Matrix']
                
                # Format header row
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Format header row
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                    cell.border = border
                
                # Format data cells
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    for cell in row:
                        cell.border = border
                        cell.alignment = center_align
                        if cell.column == 1:  # Impact Type column
                            cell.font = Font(bold=True)
                            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            # For numeric cells, use shorter width
                            if isinstance(cell.value, (int, float)):
                                max_length = max(max_length, 8)
                            else:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                        except Exception:
                            pass
                    # First column (Impact Type) can be wider, others narrower
                    if column_letter == 'A':
                        adjusted_width = min(max_length + 2, 50)
                    else:
                        adjusted_width = min(max_length + 2, 15)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Format numeric cells (counts) as integers
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=2):
                    for cell in row:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = '0'  # Integer format
                
                # Freeze first row and first column
                worksheet.freeze_panes = 'B2'
            
            year_label = f" for year {year}" if year is not None else ""
            logger.debug(f"Exported risk matrix{year_label} to {excel_file}")
            return excel_file
            
        except Exception as e:
            year_label = f" for year {year}" if year is not None else ""
            logger.error(f"Error compiling risk matrix{year_label}: {e}", exc_info=True)
            # Create empty matrix on error
            self._create_empty_risk_matrix(excel_file)
            return excel_file
    
    def _create_empty_risk_matrix(self, excel_file: Path):
        """Create an empty risk matrix template."""
        try:
            data = {'Impact Type': VALID_IMPACT_TYPES}
            for group_name in STAKEHOLDER_GROUPS.keys():
                data[group_name] = ['' for _ in VALID_IMPACT_TYPES]
            
            df = pd.DataFrame(data)
            df.to_excel(excel_file, sheet_name='Risk Matrix', index=False, engine='openpyxl')
            logger.info(f"Created empty risk matrix template at {excel_file}")
        except Exception as e:
            logger.error(f"Error creating empty risk matrix: {e}")


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description='Classify cyber events using ASD risk matrix framework',
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    
    parser.add_argument(
        '--db-path',
        type=str,
        default='instance/cyber_events.db',
        help='Path to the SQLite database (default: instance/cyber_events.db)'
    )
    
    parser.add_argument(
        '--model',
        type=str,
        default='gpt-4o',
        help='OpenAI model to use (default: gpt-4o)'
    )
    
    parser.add_argument(
        '--limit',
        type=int,
        default=5,
        help='Number of events to process (default: 5)'
    )
    
    parser.add_argument(
        '--force-reclassify',
        action='store_true',
        help='Re-classify events even if cached classification exists'
    )
    
    parser.add_argument(
        '--output-dir',
        type=str,
        default='risk_matrix',
        help='Output directory for export files (default: risk_matrix)'
    )
    
    args = parser.parse_args()
    
    try:
        # Initialize classifier
        classifier = ASDRiskClassifier(args.db_path, args.model)
        
        try:
            # Process events
            results = classifier.process_events(args.limit, args.force_reclassify)
            
            if results:
                # Export results
                csv_file, json_file, excel_files = classifier.export_results(results, args.output_dir)
                print(f"\n[SUCCESS] Classification complete!")
                print(f"   Processed {len(results)} events")
                print(f"   CSV exported to: {csv_file}")
                print(f"   JSON exported to: {json_file}")
                print(f"   Risk matrices exported:")
                for excel_file in excel_files:
                    print(f"     - {excel_file}")
            else:
                print("\n[WARNING] No events were successfully classified")
                # Still compile risk matrix from existing classifications
                output_path = Path(args.output_dir)
                output_path.mkdir(parents=True, exist_ok=True)
                excel_files = classifier.compile_risk_matrix(output_path)
                print(f"   Risk matrices compiled from existing classifications:")
                for excel_file in excel_files:
                    print(f"     - {excel_file}")

        finally:
            classifier.close()

    except Exception as e:
        logger.error(f"Fatal error: {e}", exc_info=True)
        print(f"\n[ERROR] {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()

