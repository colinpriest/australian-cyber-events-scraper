"""Tests for the data.gov.au OAIC ingester.

OAIC moved its half-yearly NDB statistics to an XLSX resource on data.gov.au
from Jul-Dec 2025 onward, so this path is the only source of new government
data. These tests cover the pure parsing/verification helpers and a
round-trip through a synthetic workbook matching the real sheet layout, so
none of them touch the network.
"""
from __future__ import annotations

import json

import pytest
from openpyxl import Workbook

from scripts.build_static_dashboard import calculate_stats_from_distribution
from scripts.oaic.oaic_datagov_scraper import (
    _normalize_range_label,
    _normalize_time_bucket,
    build_record,
    merge_into_existing,
    parse_period_from_name,
    parse_workbook,
    verify_record,
)


# --------------------------------------------------------------------------
# Period parsing
# --------------------------------------------------------------------------

@pytest.mark.parametrize(
    "name, expected",
    [
        ("NDB Data 1 July 2025 to 31 Dec 2025", (2025, "H2", 7, 12)),
        ("NDB Data 1 January 2026 to 30 June 2026", (2026, "H1", 1, 6)),
        ("NDB Data 1 Jul 2024 to 31 December 2024", (2024, "H2", 7, 12)),
    ],
)
def test_parse_period_from_name(name, expected):
    assert parse_period_from_name(name) == expected


def test_parse_period_from_name_rejects_unparseable():
    assert parse_period_from_name("Glossary and notes") is None


# --------------------------------------------------------------------------
# Label normalisation
# --------------------------------------------------------------------------

@pytest.mark.parametrize(
    "raw, expected",
    [
        ("2 - 10", "2-10"),
        ("101 - 1,000", "101-1,000"),
        ("10,000,001 or more", "10,000,001+"),
        ("Unknown", "Unknown"),
        ("1", "1"),
        # OAIC typo in the Jul-Dec 2025 workbook: every other bucket starts
        # at n+1, so 250,000 must be read as 250,001.
        ("250,000 - 500,000", "250,001-500,000"),
    ],
)
def test_normalize_range_label(raw, expected):
    assert _normalize_range_label(raw) == expected


@pytest.mark.parametrize(
    "raw, expected",
    [
        ("≤10", "<= 10 days"),
        ("≤ 10 ", "<= 10 days"),
        ("11-20", "11-20 days"),
        ("21-30", "21-30 days"),
        (">30", "> 30 days"),
        ("nonsense", None),
    ],
)
def test_normalize_time_bucket(raw, expected):
    assert _normalize_time_bucket(raw) == expected


def test_new_top_bucket_is_priced_by_dashboard():
    """The 10,000,001+ bucket must not be silently dropped from the stats.

    calculate_stats_from_distribution skips any bucket whose label is absent
    from its midpoint/bounds tables, so an unrecognised label would quietly
    vanish from the average rather than raise.
    """
    with_top = calculate_stats_from_distribution(
        [{"range": "1", "count": 1}, {"range": "10,000,001+", "count": 1}]
    )
    without_top = calculate_stats_from_distribution([{"range": "1", "count": 1}])
    assert with_top[0] is not None
    assert with_top[0] > without_top[0]


# --------------------------------------------------------------------------
# Workbook round-trip
# --------------------------------------------------------------------------

def _build_synthetic_workbook(path):
    """Create a workbook mirroring the real OAIC sheet layout, in miniature."""
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("NDB by month")
    ws.append(["Notifiable data breaches, by month and source, July-December 2025"])
    ws.append(["Month/Source", "Count"])
    for month in ("July", "August"):
        ws.append([month, 10])
        ws.append(["Human error", 4])
        ws.append(["Malicious or criminal attack", 5])
        ws.append(["System fault", 1])
    ws.append(["Grand Total", 20])

    ws = wb.create_sheet("Source of breach")
    ws.append(["Specific source of breaches, July-December 2025"])
    ws.append(["Source", "Count"])
    ws.append(["Human error", 8])
    ws.append(["PI sent to wrong recipient (email)", 8])
    ws.append(["Malicious or criminal attack", 10])
    ws.append(["Cyber incident", 7])
    ws.append(["Social engineering / impersonation", 3])
    ws.append(["System fault", 2])
    ws.append(["Unintended access", 2])
    # The real workbook repeats its final row; it must not be double counted.
    ws.append(["Unintended access", 2])
    ws.append(["Notes: shows specific source of breaches only"])

    ws = wb.create_sheet("Individuals affected")
    ws.append(["Number of individuals affected by breaches, July-December 2025"])
    ws.append(["Number of individuals world-wide affected by breaches"])
    ws.append(["Range", "Count"])
    ws.append(["1", 12])
    ws.append(["2 - 10", 7])
    ws.append(["10,000,001 or more", 1])
    ws.append(["Grand Total", 20])
    ws.append(["Large-scale data breaches affecting Australians"])
    ws.append(["Range", "Count"])
    ws.append(["100,001 - 250,000", 1])

    ws = wb.create_sheet("Personal information")
    ws.append(["Kinds of personal information involved, July-December 2025"])
    ws.append(["Kind of personal information", "Count"])
    ws.append(["Contact information", 15])
    ws.append(["Tax File Numbers", 3])

    ws = wb.create_sheet("Top 5 sectors by source")
    ws.append(["Top 5 sectors by source of breaches, July-December 2025"])
    ws.append(["Top sectors by source of breaches", "Count"])
    ws.append(["Health service providers", 11])
    ws.append(["Human error", 5])
    ws.append(["Malicious or criminal attack", 5])
    ws.append(["System fault", 1])
    ws.append(["Finance (incl. superannuation)", 9])
    ws.append(["Human error", 3])
    ws.append(["Malicious or criminal attack", 5])
    ws.append(["System fault", 1])

    ws = wb.create_sheet("Time to identify by source")
    ws.append(["Time taken to identify breach, July-December 2025"])
    ws.append(["Source and time taken (days)", "Count"])
    ws.append(["Human error"])
    ws.append(["≤10", 6])
    ws.append([">30", 2])
    ws.append(["Malicious or criminal attack"])
    ws.append(["≤10", 9])
    ws.append([">30", 3])
    ws.append(["Note: Excludes out of range date values"])

    wb.save(path)
    return path


@pytest.fixture()
def synthetic_record(tmp_path):
    path = _build_synthetic_workbook(tmp_path / "ndb.xlsx")
    sections = parse_workbook(path)
    return build_record(sections, 2025, "H2", 7, 12, "https://example.invalid/ndb.xlsx")


def test_record_headline_figures(synthetic_record):
    record = synthetic_record
    assert record["year"] == 2025
    assert record["period"] == "H2"
    assert record["total_notifications"] == 20
    assert record["human_error"] == 8
    assert record["malicious_attacks"] == 10
    assert record["system_faults"] == 2
    assert record["source"] == "data.gov.au"


def test_duplicate_workbook_row_not_double_counted(synthetic_record):
    """System fault sub-sources must total the category count, not twice it."""
    system_fault = synthetic_record["breach_sub_sources"]["System fault"]
    assert system_fault == {"Unintended access": 2}


def test_cyber_incidents_follows_existing_series_convention(synthetic_record):
    """Every dashboard-scraped period from 2023 H1 on sets cyber_incidents_total
    to the malicious-attack count. Diverging here would put a definitional
    break in the middle of the dashboard's trend line, so the narrower
    "Cyber incident" sub-source is carried separately instead.
    """
    assert synthetic_record["cyber_incidents_total"] == synthetic_record["malicious_attacks"]
    assert synthetic_record["cyber_incident_only"] == 7


def test_unpublished_attack_methods_are_null_not_zero(synthetic_record):
    """data.gov.au omits the per-method split; zeros would read as real data."""
    for field in ("phishing", "ransomware", "hacking", "brute_force",
                  "malware", "compromised_credentials"):
        assert synthetic_record[field] is None


def test_top_sectors_and_matrix(synthetic_record):
    assert synthetic_record["top_sectors"] == [
        {"sector": "Health service providers", "notifications": 11},
        {"sector": "Finance (incl. superannuation)", "notifications": 9},
    ]
    assert synthetic_record["sector_by_source"][0] == {
        "sector": "Health service providers",
        "human_error": 5,
        "malicious_or_criminal": 5,
        "system_fault": 1,
    }


def test_time_buckets_aggregate_across_sources(synthetic_record):
    buckets = {e["bucket"]: e["count"] for e in synthetic_record["time_to_identify"]}
    assert buckets == {"<= 10 days": 15, "> 30 days": 5}


def test_verify_record_passes_on_consistent_data(synthetic_record):
    assert verify_record(synthetic_record) == []


def test_verify_record_flags_inconsistent_totals(synthetic_record):
    broken = dict(synthetic_record)
    broken["human_error"] = 99
    problems = verify_record(broken)
    assert any("source breakdown" in p for p in problems)


# --------------------------------------------------------------------------
# Merge behaviour
# --------------------------------------------------------------------------

def test_merge_replaces_matching_period_and_keeps_others(tmp_path):
    existing = [
        {"year": 2025, "period": "H1", "total_notifications": 532},
        {"year": 2025, "period": "H2", "total_notifications": 1},
    ]
    path = tmp_path / "existing.json"
    path.write_text(json.dumps(existing), encoding="utf-8")

    merged = merge_into_existing(
        [{"year": 2025, "period": "H2", "total_notifications": 670}], path
    )

    assert [(r["year"], r["period"]) for r in merged] == [(2025, "H1"), (2025, "H2")]
    assert merged[0]["total_notifications"] == 532
    assert merged[1]["total_notifications"] == 670
