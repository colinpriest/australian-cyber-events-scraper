#!/usr/bin/env python3
"""
Export deduplicated cyber events to Excel with LLM-summarized descriptions.

This script:
1. Reads all active deduplicated events from the database
2. Collects all associated text from EnrichedEvents and RawEvents
3. Uses OpenAI GPT to summarize each event
4. Creates anonymized versions with entity names replaced by generic terms
5. Outputs to Excel with columns:
   - Event Date
   - Event Title
   - Event Description (LLM-summarized)
   - Anonymised Description (entity names removed)
   - Records Affected
   - Entity Type (industry category)
   - Attack Type

Usage:
    python export_events_excel.py
    python export_events_excel.py --output events_export.xlsx
    python export_events_excel.py --limit 100  # Export only first 100 events
    python export_events_excel.py --no-llm     # Skip LLM summarization (use raw text)
"""

import argparse
import json
import os
import sqlite3
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from dotenv import load_dotenv
from openai import OpenAI
import pandas as pd
from tqdm import tqdm

# Load environment variables
load_dotenv()


def get_openai_client() -> Optional[OpenAI]:
    """Initialize OpenAI client if API key is available."""
    api_key = os.getenv('OPENAI_API_KEY')
    if not api_key:
        print("Warning: OPENAI_API_KEY not found. LLM summarization will be disabled.")
        return None
    return OpenAI(api_key=api_key)


def get_event_source_text(cursor: sqlite3.Cursor, deduplicated_event_id: str, event_title: str = None) -> str:
    """Collect all source text for a deduplicated event.

    Uses multiple strategies to find content:
    1. Direct linkage via EventDeduplicationMap
    2. Title matching to EnrichedEvents (fallback for broken linkages)
    """
    texts = []

    # Get the deduplicated event's own description
    cursor.execute("""
        SELECT title, description, summary
        FROM DeduplicatedEvents
        WHERE deduplicated_event_id = ?
    """, (deduplicated_event_id,))
    row = cursor.fetchone()
    dedup_title = None
    if row:
        dedup_title = row[0]
        if row[1]:
            texts.append(f"Description: {row[1]}")
        if row[2]:
            texts.append(f"Summary: {row[2]}")

    # Strategy 1: Try direct linkage via EventDeduplicationMap
    cursor.execute("""
        SELECT DISTINCT ee.title, ee.description, ee.summary
        FROM EventDeduplicationMap edm
        JOIN EnrichedEvents ee ON edm.enriched_event_id = ee.enriched_event_id
        WHERE edm.deduplicated_event_id = ?
    """, (deduplicated_event_id,))
    enriched_rows = cursor.fetchall()

    # Strategy 2: If no direct linkage, try title matching
    if not enriched_rows and (dedup_title or event_title):
        title_to_match = event_title or dedup_title
        # Use first 30 chars for matching to handle truncated titles
        title_prefix = title_to_match[:30] if title_to_match else None
        if title_prefix:
            cursor.execute("""
                SELECT DISTINCT ee.title, ee.description, ee.summary
                FROM EnrichedEvents ee
                WHERE ee.title LIKE ?
                LIMIT 3
            """, (f"{title_prefix}%",))
            enriched_rows = cursor.fetchall()

    for row in enriched_rows:
        if row[1] and row[1] not in str(texts):
            texts.append(f"Event Description: {row[1]}")
        if row[2] and row[2] not in str(texts):
            texts.append(f"Event Summary: {row[2]}")

    # Get raw event content (the actual scraped article text)
    # Strategy 1: Via EventDeduplicationMap -> EnrichedEvents -> RawEvents
    cursor.execute("""
        SELECT DISTINCT re.raw_title, re.raw_description,
               SUBSTR(re.raw_content, 1, 10000) as content_preview
        FROM EventDeduplicationMap edm
        JOIN EnrichedEvents ee ON edm.enriched_event_id = ee.enriched_event_id
        JOIN RawEvents re ON ee.raw_event_id = re.raw_event_id
        WHERE edm.deduplicated_event_id = ?
        LIMIT 3
    """, (deduplicated_event_id,))
    raw_rows = cursor.fetchall()

    # Strategy 2: If no direct linkage, find raw content via title matching
    if not raw_rows and (dedup_title or event_title):
        title_to_match = event_title or dedup_title
        title_prefix = title_to_match[:30] if title_to_match else None
        if title_prefix:
            cursor.execute("""
                SELECT DISTINCT re.raw_title, re.raw_description,
                       SUBSTR(re.raw_content, 1, 10000) as content_preview
                FROM EnrichedEvents ee
                JOIN RawEvents re ON ee.raw_event_id = re.raw_event_id
                WHERE ee.title LIKE ?
                LIMIT 3
            """, (f"{title_prefix}%",))
            raw_rows = cursor.fetchall()

    for row in raw_rows:
        if row[2] and len(row[2]) > 100:
            texts.append(f"Article Content:\n{row[2]}")

    return "\n\n".join(texts)


def summarize_with_llm(client: OpenAI, text: str, max_words: int = 500) -> str:
    """Use OpenAI to summarize text to max_words."""
    if not text or len(text.strip()) < 50:
        return text or ""

    # Truncate input if too long (to stay within token limits)
    if len(text) > 30000:
        text = text[:30000] + "\n\n[Text truncated...]"

    try:
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {
                    "role": "system",
                    "content": f"""You are a cybersecurity analyst summarizing data breach events.
Create a comprehensive summary of the cyber security incident described below.
The summary should be factual, professional, and include:
- What happened (type of attack/breach)
- Who was affected (organization and individuals)
- What data was compromised (be specific about data types)
- When it occurred (if known)
- Scale of impact (number of records/individuals affected)
- Any response or remediation mentioned

Write up to {max_words} words. Be thorough and include all relevant details."""
                },
                {
                    "role": "user",
                    "content": f"Summarize this cyber security event:\n\n{text}"
                }
            ],
            max_tokens=2000,
            temperature=0.3
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        print(f"Warning: LLM summarization failed: {e}")
        # Return truncated raw text as fallback
        words = text.split()
        if len(words) > max_words:
            return " ".join(words[:max_words]) + "..."
        return text


def anonymize_with_llm(client: OpenAI, text: str, industry: str = None) -> str:
    """Use OpenAI to anonymize entity names in the description."""
    if not text or len(text.strip()) < 20:
        return text or ""

    industry_hint = f"The organization operates in the {industry} sector." if industry else ""

    try:
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {
                    "role": "system",
                    "content": f"""You are a cybersecurity analyst anonymizing incident reports.
Replace all specific entity names with generic industry terms while preserving the incident details.

Rules:
- Replace company/organization names with generic descriptions based on their industry
  (e.g., "Medibank" → "a major health insurer", "Optus" → "a telecommunications provider")
- Replace person names with generic roles (e.g., "John Smith, CEO" → "the company's CEO")
- Keep all technical details, dates, numbers, and incident specifics intact
- Preserve the meaning and severity of the incident
- Do not add any new information
{industry_hint}

Return ONLY the anonymized text, no explanations."""
                },
                {
                    "role": "user",
                    "content": f"Anonymize this incident description:\n\n{text}"
                }
            ],
            max_tokens=2000,
            temperature=0.2
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        print(f"Warning: LLM anonymization failed: {e}")
        return text  # Return original if anonymization fails


def truncate_text(text: str, max_words: int = 500) -> str:
    """Simple truncation without LLM."""
    if not text:
        return ""
    words = text.split()
    if len(words) > max_words:
        return " ".join(words[:max_words]) + "..."
    return text


def process_single_event(
    client: OpenAI,
    event_data: Dict,
    source_text: str,
    max_words: int,
    use_llm: bool
) -> Dict:
    """Process a single event with LLM summarization and anonymization."""
    industry = event_data['industry']

    # Summarize or truncate
    if use_llm and client and len(source_text) > 100:
        description = summarize_with_llm(client, source_text, max_words)
        # Create anonymized version
        anonymized = anonymize_with_llm(client, description, industry)
    else:
        description = truncate_text(source_text, max_words)
        anonymized = description

    return {
        'Event Date': event_data['event_date'],
        'Event Title': event_data['title'] or 'Untitled Event',
        'Event Description': description,
        'Anonymised Description': anonymized,
        'Records Affected': event_data['records_affected'],
        'Entity Type': industry,
        'Attack Type': event_data['event_type']
    }


def export_events_to_excel(
    db_path: str = 'instance/cyber_events.db',
    output_path: str = None,
    limit: int = None,
    use_llm: bool = True,
    max_words: int = 500,
    max_workers: int = 10
) -> str:
    """Export deduplicated events to Excel file with parallel LLM processing."""

    if output_path is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_path = f'cyber_events_export_{timestamp}.xlsx'

    # Initialize OpenAI client if using LLM
    client = None
    if use_llm:
        client = get_openai_client()
        if client is None:
            print("Falling back to simple text truncation (no LLM).")
            use_llm = False

    # Connect to database
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    # Get all active deduplicated events
    query = """
        SELECT
            deduplicated_event_id,
            title,
            event_date,
            event_type,
            victim_organization_industry,
            records_affected,
            description,
            summary
        FROM DeduplicatedEvents
        WHERE status = 'Active'
        ORDER BY event_date DESC
    """
    if limit:
        query += f" LIMIT {limit}"

    cursor.execute(query)
    events = cursor.fetchall()

    print(f"Processing {len(events)} events...")

    # Phase 1: Collect all source texts (sequential - uses DB)
    print("Phase 1: Gathering source texts from database...")
    event_data_list = []
    for event in tqdm(events, desc="Reading events"):
        event_id = event['deduplicated_event_id']
        event_title = event['title']

        # Clean up event type
        event_type = event['event_type'] or 'Unknown'
        if '.' in event_type:
            event_type = event_type.split('.')[-1].replace('_', ' ').title()

        source_text = get_event_source_text(cursor, event_id, event_title)

        event_data_list.append({
            'event_data': {
                'event_date': event['event_date'],
                'title': event['title'],
                'records_affected': event['records_affected'],
                'industry': event['victim_organization_industry'] or 'Unknown',
                'event_type': event_type
            },
            'source_text': source_text
        })

    conn.close()

    # Phase 2: Process with LLM in parallel
    print(f"Phase 2: Processing with LLM ({max_workers} parallel workers)...")
    rows = [None] * len(event_data_list)  # Pre-allocate to maintain order

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        # Submit all tasks
        future_to_idx = {
            executor.submit(
                process_single_event,
                client,
                item['event_data'],
                item['source_text'],
                max_words,
                use_llm
            ): idx
            for idx, item in enumerate(event_data_list)
        }

        # Collect results with progress bar
        for future in tqdm(as_completed(future_to_idx), total=len(future_to_idx), desc="LLM processing"):
            idx = future_to_idx[future]
            try:
                rows[idx] = future.result()
            except Exception as e:
                print(f"Error processing event {idx}: {e}")
                # Fallback to basic data
                item = event_data_list[idx]
                rows[idx] = {
                    'Event Date': item['event_data']['event_date'],
                    'Event Title': item['event_data']['title'] or 'Untitled Event',
                    'Event Description': truncate_text(item['source_text'], max_words),
                    'Anonymised Description': truncate_text(item['source_text'], max_words),
                    'Records Affected': item['event_data']['records_affected'],
                    'Entity Type': item['event_data']['industry'],
                    'Attack Type': item['event_data']['event_type']
                }

    # Create DataFrame and export to Excel
    df = pd.DataFrame(rows)

    # Write to Excel with formatting
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Cyber Events')

        # Get worksheet and apply formatting
        worksheet = writer.sheets['Cyber Events']
        from openpyxl.styles import Alignment

        # Define column widths - wider for description columns
        column_widths = {
            'A': 12,   # Event Date
            'B': 40,   # Event Title
            'C': 80,   # Event Description
            'D': 80,   # Anonymised Description
            'E': 15,   # Records Affected
            'F': 25,   # Entity Type
            'G': 20,   # Attack Type
        }

        # Apply column widths
        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width

        # Enable text wrapping for description columns (C and D)
        # and set appropriate row heights
        for row_num in range(2, len(df) + 2):  # Start from row 2 (after header)
            # Event Description (column C)
            cell_c = worksheet.cell(row=row_num, column=3)
            cell_c.alignment = Alignment(wrap_text=True, vertical='top')

            # Anonymised Description (column D)
            cell_d = worksheet.cell(row=row_num, column=4)
            cell_d.alignment = Alignment(wrap_text=True, vertical='top')

            # Calculate row height based on content length
            desc_len = len(str(cell_c.value or ''))
            # Approximate: 80 chars per line at width 80, ~15 points per line
            estimated_lines = max(desc_len // 80, 1)
            row_height = min(estimated_lines * 15, 400)  # Cap at 400 points
            worksheet.row_dimensions[row_num].height = max(row_height, 30)

        # Format header row
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    print(f"\nExported {len(rows)} events to: {output_path}")
    return output_path


def main():
    parser = argparse.ArgumentParser(
        description='Export deduplicated cyber events to Excel with LLM-summarized descriptions.'
    )
    parser.add_argument(
        '--output', '-o',
        type=str,
        default=None,
        help='Output Excel file path (default: cyber_events_export_<timestamp>.xlsx)'
    )
    parser.add_argument(
        '--db-path',
        type=str,
        default='instance/cyber_events.db',
        help='Path to the SQLite database (default: instance/cyber_events.db)'
    )
    parser.add_argument(
        '--limit', '-l',
        type=int,
        default=None,
        help='Limit number of events to export (default: all)'
    )
    parser.add_argument(
        '--no-llm',
        action='store_true',
        help='Skip LLM summarization, use simple text truncation instead'
    )
    parser.add_argument(
        '--max-words', '-w',
        type=int,
        default=500,
        help='Maximum words for event description (default: 500)'
    )
    parser.add_argument(
        '--workers',
        type=int,
        default=10,
        help='Number of parallel workers for LLM processing (default: 10)'
    )

    args = parser.parse_args()

    # Check if openpyxl is available
    try:
        import openpyxl
    except ImportError:
        print("Error: openpyxl is required for Excel export.")
        print("Install with: pip install openpyxl")
        sys.exit(1)

    export_events_to_excel(
        db_path=args.db_path,
        output_path=args.output,
        limit=args.limit,
        use_llm=not args.no_llm,
        max_words=args.max_words,
        max_workers=args.workers
    )


if __name__ == '__main__':
    main()
