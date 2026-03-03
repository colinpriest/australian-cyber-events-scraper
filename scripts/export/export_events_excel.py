#!/usr/bin/env python3
"""
Export deduplicated cyber events to Excel with LLM-summarized descriptions.

This script:
1. Reads all active deduplicated events from the database
2. Collects all associated text from EnrichedEvents and RawEvents
3. Uses OpenAI GPT to summarize each event
4. Creates anonymized versions with entity names replaced by generic terms
5. Outputs to Excel with columns:
   - Event Date
   - Event Title
   - Event Description (LLM-summarized)
   - Anonymised Description (entity names removed, no dates/years, no title)
   - Records Affected
   - Entity Type (industry category)
   - Attack Type

Usage:
    python export_events_excel.py
    python export_events_excel.py --output events_export.xlsx
    python export_events_excel.py --limit 100  # Export only first 100 events
    python export_events_excel.py --no-llm     # Skip LLM summarization (use raw text)
    python export_events_excel.py --exclude-unknown-records  # Exclude events with unknown records affected
"""

from __future__ import annotations

import argparse
import json
import os
import sqlite3
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

from dotenv import load_dotenv
from openai import OpenAI
import pandas as pd
from tqdm import tqdm

# Load environment variables
load_dotenv()


def get_openai_client() -> Optional[OpenAI]:
    """Initialize OpenAI client if API key is available."""
    api_key = os.getenv('OPENAI_API_KEY')
    if not api_key:
        print("Warning: OPENAI_API_KEY not found. LLM summarization will be disabled.")
        return None
    return OpenAI(api_key=api_key)


def get_event_source_text(cursor: sqlite3.Cursor, deduplicated_event_id: str, event_title: str = None) -> str:
    """Collect all source text for a deduplicated event.

    Uses multiple strategies to find content:
    1. Direct linkage via EventDeduplicationMap
    2. Title matching to EnrichedEvents (fallback for broken linkages)
    """
    texts = []

    # Get the deduplicated event's own description
    cursor.execute("""
        SELECT title, description, summary
        FROM DeduplicatedEvents
        WHERE deduplicated_event_id = ?
    """, (deduplicated_event_id,))
    row = cursor.fetchone()
    dedup_title = None
    if row:
        dedup_title = row[0]
        if row[1]:
            texts.append(f"Description: {row[1]}")
        if row[2]:
            texts.append(f"Summary: {row[2]}")

    # Strategy 1: Try direct linkage via EventDeduplicationMap
    cursor.execute("""
        SELECT DISTINCT ee.title, ee.description, ee.summary
        FROM EventDeduplicationMap edm
        JOIN EnrichedEvents ee ON edm.enriched_event_id = ee.enriched_event_id
        WHERE edm.deduplicated_event_id = ?
    """, (deduplicated_event_id,))
    enriched_rows = cursor.fetchall()

    # Strategy 2: If no direct linkage, try title matching
    if not enriched_rows and (dedup_title or event_title):
        title_to_match = event_title or dedup_title
        # Use first 30 chars for matching to handle truncated titles
        title_prefix = title_to_match[:30] if title_to_match else None
        if title_prefix:
            cursor.execute("""
                SELECT DISTINCT ee.title, ee.description, ee.summary
                FROM EnrichedEvents ee
                WHERE ee.title LIKE ?
                LIMIT 3
            """, (f"{title_prefix}%",))
            enriched_rows = cursor.fetchall()

    for row in enriched_rows:
        if row[1] and row[1] not in str(texts):
            texts.append(f"Event Description: {row[1]}")
        if row[2] and row[2] not in str(texts):
            texts.append(f"Event Summary: {row[2]}")

    # Get raw event content (the actual scraped article text)
    # Strategy 1: Via EventDeduplicationMap -> EnrichedEvents -> RawEvents
    cursor.execute("""
        SELECT DISTINCT re.raw_title, re.raw_description,
               SUBSTR(re.raw_content, 1, 10000) as content_preview
        FROM EventDeduplicationMap edm
        JOIN EnrichedEvents ee ON edm.enriched_event_id = ee.enriched_event_id
        JOIN RawEvents re ON ee.raw_event_id = re.raw_event_id
        WHERE edm.deduplicated_event_id = ?
        LIMIT 3
    """, (deduplicated_event_id,))
    raw_rows = cursor.fetchall()

    # Strategy 2: If no direct linkage, find raw content via title matching
    if not raw_rows and (dedup_title or event_title):
        title_to_match = event_title or dedup_title
        title_prefix = title_to_match[:30] if title_to_match else None
        if title_prefix:
            cursor.execute("""
                SELECT DISTINCT re.raw_title, re.raw_description,
                       SUBSTR(re.raw_content, 1, 10000) as content_preview
                FROM EnrichedEvents ee
                JOIN RawEvents re ON ee.raw_event_id = re.raw_event_id
                WHERE ee.title LIKE ?
                LIMIT 3
            """, (f"{title_prefix}%",))
            raw_rows = cursor.fetchall()

    for row in raw_rows:
        if row[2] and len(row[2]) > 100:
            texts.append(f"Article Content:\n{row[2]}")

    return "\n\n".join(texts)


def summarize_with_llm(client: OpenAI, text: str, max_words: int = 500) -> str:
    """Use OpenAI to summarize text to max_words."""
    if not text or len(text.strip()) < 50:
        return text or ""

    # Truncate input if too long (to stay within token limits)
    if len(text) > 30000:
        text = text[:30000] + "\n\n[Text truncated...]"

    try:
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {
                    "role": "system",
                    "content": f"""You are a cybersecurity analyst summarizing data breach events.
Create a comprehensive summary of the cyber security incident described below.
The summary should be factual, professional, and include:
- What happened (type of attack/breach)
- Who was affected (organization and individuals)
- What data was compromised (be specific about data types)
- When it occurred (if known)
- Scale of impact (number of records/individuals affected)
- Any response or remediation mentioned

Write up to {max_words} words. Be thorough and include all relevant details."""
                },
                {
                    "role": "user",
                    "content": f"Summarize this cyber security event:\n\n{text}"
                }
            ],
            max_tokens=2000,
            temperature=0.3
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        print(f"Warning: LLM summarization failed: {e}")
        # Return truncated raw text as fallback
        words = text.split()
        if len(words) > max_words:
            return " ".join(words[:max_words]) + "..."
        return text


def anonymize_with_llm(client: OpenAI, text: str, industry: str = None, title: str = None, all_entity_names: List[str] = None) -> str:
    """Use OpenAI to anonymize entity names in the description, removing dates and titles."""
    if not text or len(text.strip()) < 20:
        return text or ""

    industry_hint = f"The victim organization operates in the {industry} sector." if industry and industry != "Unknown" else ""

    # Build list of known entities to help the LLM
    entity_hint = ""
    if all_entity_names and len(all_entity_names) > 0:
        # Include up to 50 entity names as examples
        sample_entities = all_entity_names[:50]
        entity_hint = f"\n\nKnown entity names that MUST be anonymized if present: {', '.join(sample_entities)}"

    try:
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {
                    "role": "system",
                    "content": f"""You are a cybersecurity analyst creating a fully anonymized incident report.
Your task is to remove ALL identifying information while preserving the incident details.

CRITICAL RULES - You MUST follow ALL of these:

1. REMOVE ALL ENTITY NAMES:
   - Replace ALL company/organization names with generic descriptions based on their industry
     (e.g., "Medibank" → "a major health insurer", "Optus" → "a telecommunications provider",
      "Commonwealth Bank" → "a major bank", "Woolworths" → "a major retailer")
   - Replace ALL person names with generic roles (e.g., "John Smith, CEO" → "the CEO")
   - Replace ALL threat actor/hacker group names with "threat actors" or "attackers"
   - Be thorough - catch ALL organization names including subsidiaries, partners, vendors

2. REMOVE ALL DATES AND TIME REFERENCES:
   - Remove ALL specific dates (e.g., "January 15, 2024", "15/01/2024", "2024-01-15")
   - Remove ALL year references (e.g., "in 2024", "during 2023", "since 2022")
   - Remove ALL month references with years (e.g., "March 2024", "Q1 2024")
   - Replace with generic terms if needed for context (e.g., "recently", "the incident")

3. REMOVE ANY TITLE AT THE START:
   - If the text begins with a title or headline, remove it completely
   - Start directly with the incident description

4. PRESERVE:
   - Technical attack details and methods
   - Types of data compromised
   - Number of records/individuals affected (keep the numbers)
   - Impact and consequences
   - Response actions taken (anonymized)

{industry_hint}{entity_hint}

Return ONLY the fully anonymized text with no dates, no entity names, and no title. Do not include any explanations."""
                },
                {
                    "role": "user",
                    "content": f"Fully anonymize this incident description (remove all entity names, dates, years, and any title):\n\n{text}"
                }
            ],
            max_tokens=2000,
            temperature=0.2
        )
        return response.choices[0].message.content.strip()
    except Exception as e:
        print(f"Warning: LLM anonymization failed: {e}")
        return text  # Return original if anonymization fails


def truncate_text(text: str, max_words: int = 500) -> str:
    """Simple truncation without LLM."""
    if not text:
        return ""
    words = text.split()
    if len(words) > max_words:
        return " ".join(words[:max_words]) + "..."
    return text


def process_single_event(
    client: OpenAI,
    event_data: Dict,
    source_text: str,
    max_words: int,
    use_llm: bool,
    all_entity_names: List[str] = None
) -> Dict:
    """Process a single event with LLM summarization and anonymization."""
    industry = event_data['industry']
    title = event_data['title']

    # Summarize or truncate
    if use_llm and client and len(source_text) > 100:
        description = summarize_with_llm(client, source_text, max_words)
        # Create anonymized version (with title, entity names for thorough anonymization)
        anonymized = anonymize_with_llm(client, description, industry, title, all_entity_names)
    else:
        description = truncate_text(source_text, max_words)
        anonymized = description

    return {
        'Event Date': event_data['event_date'],
        'Event Title': title or 'Untitled Event',
        'Event Description': description,
        'Anonymised Description': anonymized,
        'Records Affected': event_data['records_affected'],
        'Entity Type': industry,
        'Attack Type': event_data['event_type']
    }


def get_all_entity_names(cursor: sqlite3.Cursor) -> List[str]:
    """Get all entity names from the database for thorough anonymization."""
    entity_names = set()

    # Get entity names from EntitiesV2
    try:
        cursor.execute("SELECT DISTINCT entity_name FROM EntitiesV2 WHERE entity_name IS NOT NULL")
        for row in cursor.fetchall():
            if row[0]:
                entity_names.add(row[0].strip())
    except sqlite3.Error:
        pass

    # Get victim organization names from DeduplicatedEvents
    try:
        cursor.execute("SELECT DISTINCT victim_organization_name FROM DeduplicatedEvents WHERE victim_organization_name IS NOT NULL")
        for row in cursor.fetchall():
            if row[0]:
                entity_names.add(row[0].strip())
    except sqlite3.Error:
        pass

    # Get attacking entity names from DeduplicatedEvents
    try:
        cursor.execute("SELECT DISTINCT attacking_entity_name FROM DeduplicatedEvents WHERE attacking_entity_name IS NOT NULL")
        for row in cursor.fetchall():
            if row[0]:
                entity_names.add(row[0].strip())
    except sqlite3.Error:
        pass

    # Sort by length descending to replace longer names first
    return sorted(list(entity_names), key=len, reverse=True)


def export_events_to_excel(
    db_path: str = 'instance/cyber_events.db',
    output_path: str = None,
    limit: int = None,
    use_llm: bool = True,
    max_words: int = 500,
    max_workers: int = 10,
    exclude_unknown_records: bool = False
) -> str:
    """Export deduplicated events to Excel file with parallel LLM processing."""

    if output_path is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_path = f'cyber_events_export_{timestamp}.xlsx'

    # Initialize OpenAI client if using LLM
    client = None
    if use_llm:
        client = get_openai_client()
        if client is None:
            print("Falling back to simple text truncation (no LLM).")
            use_llm = False

    # Check database exists before connecting
    if not Path(db_path).exists():
        print(f"Error: Database not found at {db_path}")
        return output_path

    # Connect to database using context manager to ensure the connection is always closed
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        # Get all entity names for thorough anonymization
        all_entity_names = get_all_entity_names(cursor) if use_llm else []
        print(f"Loaded {len(all_entity_names)} entity names for anonymization")

        # Build query with optional filter for unknown records
        query = """
            SELECT
                deduplicated_event_id,
                title,
                event_date,
                event_type,
                victim_organization_industry,
                records_affected,
                description,
                summary
            FROM DeduplicatedEvents
            WHERE status = 'Active'
        """

        if exclude_unknown_records:
            query += " AND records_affected IS NOT NULL AND records_affected != '' AND CAST(records_affected AS TEXT) != 'Unknown'"

        query += " ORDER BY event_date DESC"

        params: tuple
        if limit:
            query += " LIMIT ?"
            params = (limit,)
        else:
            params = ()

        cursor.execute(query, params)
        events = cursor.fetchall()

        print(f"Processing {len(events)} events...")

        # Phase 1: Collect all source texts (sequential - uses DB)
        print("Phase 1: Gathering source texts from database...")
        event_data_list = []
        for event in tqdm(events, desc="Reading events"):
            event_id = event['deduplicated_event_id']
            event_title = event['title']

            # Clean up event type - use "Unknown" for unknown/null values
            event_type = event['event_type']
            if not event_type or event_type.lower() in ('unknown', 'none', '', 'null'):
                event_type = 'Unknown'
            elif '.' in event_type:
                event_type = event_type.split('.')[-1].replace('_', ' ').title()

            # Clean up industry - use "Unknown" for unknown/null values
            industry = event['victim_organization_industry']
            if not industry or industry.lower() in ('unknown', 'none', '', 'null'):
                industry = 'Unknown'

            source_text = get_event_source_text(cursor, event_id, event_title)

            event_data_list.append({
                'event_data': {
                    'event_date': event['event_date'],
                    'title': event['title'],
                    'records_affected': event['records_affected'],
                    'industry': industry,
                    'event_type': event_type
                },
                'source_text': source_text
            })

    # Phase 2: Process with LLM in parallel
    print(f"Phase 2: Processing with LLM ({max_workers} parallel workers)...")
    rows = [None] * len(event_data_list)  # Pre-allocate to maintain order

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        # Submit all tasks
        future_to_idx = {
            executor.submit(
                process_single_event,
                client,
                item['event_data'],
                item['source_text'],
                max_words,
                use_llm,
                all_entity_names
            ): idx
            for idx, item in enumerate(event_data_list)
        }

        # Collect results with progress bar
        for future in tqdm(as_completed(future_to_idx), total=len(future_to_idx), desc="LLM processing"):
            idx = future_to_idx[future]
            try:
                rows[idx] = future.result()
            except Exception as e:
                print(f"Error processing event {idx}: {e}")
                # Fallback to basic data
                item = event_data_list[idx]
                rows[idx] = {
                    'Event Date': item['event_data']['event_date'],
                    'Event Title': item['event_data']['title'] or 'Untitled Event',
                    'Event Description': truncate_text(item['source_text'], max_words),
                    'Anonymised Description': truncate_text(item['source_text'], max_words),
                    'Records Affected': item['event_data']['records_affected'],
                    'Entity Type': item['event_data']['industry'],
                    'Attack Type': item['event_data']['event_type']
                }

    # Create DataFrame and export to Excel
    df = pd.DataFrame(rows)

    # Write to Excel with formatting
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Cyber Events')

        # Get worksheet and apply formatting
        worksheet = writer.sheets['Cyber Events']
        from openpyxl.styles import Alignment

        # Define column widths - wider for description columns
        column_widths = {
            'A': 12,   # Event Date
            'B': 40,   # Event Title
            'C': 80,   # Event Description
            'D': 80,   # Anonymised Description
            'E': 15,   # Records Affected
            'F': 25,   # Entity Type
            'G': 20,   # Attack Type
        }

        # Apply column widths
        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width

        # Enable text wrapping for description columns (C and D)
        # and set appropriate row heights
        for row_num in range(2, len(df) + 2):  # Start from row 2 (after header)
            # Event Description (column C)
            cell_c = worksheet.cell(row=row_num, column=3)
            cell_c.alignment = Alignment(wrap_text=True, vertical='top')

            # Anonymised Description (column D)
            cell_d = worksheet.cell(row=row_num, column=4)
            cell_d.alignment = Alignment(wrap_text=True, vertical='top')

            # Calculate row height based on content length
            desc_len = len(str(cell_c.value or ''))
            # Approximate: 80 chars per line at width 80, ~15 points per line
            estimated_lines = max(desc_len // 80, 1)
            row_height = min(estimated_lines * 15, 400)  # Cap at 400 points
            worksheet.row_dimensions[row_num].height = max(row_height, 30)

        # Format header row
        for col_num in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    print(f"\nExported {len(rows)} events to: {output_path}")
    return output_path


def main():
    parser = argparse.ArgumentParser(
        description='Export deduplicated cyber events to Excel with LLM-summarized descriptions.'
    )
    parser.add_argument(
        '--output', '-o',
        type=str,
        default=None,
        help='Output Excel file path (default: cyber_events_export_<timestamp>.xlsx)'
    )
    parser.add_argument(
        '--db-path',
        type=str,
        default='instance/cyber_events.db',
        help='Path to the SQLite database (default: instance/cyber_events.db)'
    )
    parser.add_argument(
        '--limit', '-l',
        type=int,
        default=None,
        help='Limit number of events to export (default: all)'
    )
    parser.add_argument(
        '--no-llm',
        action='store_true',
        help='Skip LLM summarization, use simple text truncation instead'
    )
    parser.add_argument(
        '--max-words', '-w',
        type=int,
        default=500,
        help='Maximum words for event description (default: 500)'
    )
    parser.add_argument(
        '--workers',
        type=int,
        default=10,
        help='Number of parallel workers for LLM processing (default: 10)'
    )
    parser.add_argument(
        '--exclude-unknown-records',
        action='store_true',
        help='Exclude events where the number of customer records affected is unknown'
    )

    args = parser.parse_args()

    # Check if openpyxl is available
    try:
        import openpyxl
    except ImportError:
        print("Error: openpyxl is required for Excel export.")
        print("Install with: pip install openpyxl")
        sys.exit(1)

    output_file = export_events_to_excel(
        db_path=args.db_path,
        output_path=args.output,
        limit=args.limit,
        use_llm=not args.no_llm,
        max_words=args.max_words,
        max_workers=args.workers,
        exclude_unknown_records=args.exclude_unknown_records
    )
    print(f"\nOutput file: {Path(output_file).resolve()}")


if __name__ == '__main__':
    main()
