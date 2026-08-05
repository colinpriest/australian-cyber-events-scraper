"""Ingest OAIC Notifiable Data Breaches statistics published on data.gov.au.

From the July-December 2025 period onward the OAIC stopped publishing its
half-yearly NDB statistics through the two channels this project already
scrapes:

  * the Power BI dashboard (``OAIC_dashboard_scraper.py``) still tops out at
    Jan-Jun 2025 and states "Updates to this dashboard are forthcoming"; and
  * the PDF publications page (``oaic_data_scraper.py``) stops at Jul-Dec 2024.

The data itself is published instead as an XLSX resource on data.gov.au under
the "Notifiable Data Breaches (NDB) scheme" dataset. This module discovers
those resources through the CKAN API, parses the workbook, and emits records
in the same schema as the dashboard/PDF scrapers so that
``build_static_dashboard.py`` can consume them unchanged.

Resources are discovered dynamically rather than hard-coded, so a future
Jan-Jun 2026 release is picked up automatically.

Usage:
    python scripts/oaic/oaic_datagov_scraper.py
    python scripts/oaic/oaic_datagov_scraper.py --existing-data oaic_cyber_statistics_X.json
    python scripts/oaic/oaic_datagov_scraper.py --dry-run
"""

from __future__ import annotations

import argparse
import json
import logging
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import requests
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

CKAN_PACKAGE_URL = "https://data.gov.au/data/api/3/action/package_show"
DATASET_ID = "notifiable-data-breaches-ndb-scheme"
DASHBOARD_URL = (
    "https://www.oaic.gov.au/privacy/notifiable-data-breaches/"
    "notifiable-data-breach-statistics-dashboard"
)
USER_AGENT = "Mozilla/5.0 (compatible; australian-cyber-events-scraper/1.0)"

MONTH_NAMES: Tuple[str, ...] = (
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
)
MONTH_TO_NUM: Dict[str, int] = {m.lower(): i for i, m in enumerate(MONTH_NAMES, start=1)}

# The five source-of-breach labels OAIC uses as sub-rows beneath a month or
# sector heading. Used to tell heading rows apart from their breakdown rows.
SOURCE_LABELS = frozenset({
    "Currently unknown",
    "Human error",
    "Malicious or criminal attack",
    "Other",
    "System fault",
})

# Top-level categories on the "Source of breach" sheet; everything else on
# that sheet is a specific sub-source belonging to the preceding category.
SOURCE_CATEGORIES = frozenset({
    "Human error",
    "Malicious or criminal attack",
    "System fault",
})

PERSONAL_INFO_FIELD_MAP: Dict[str, str] = {
    "contact information": "contact_information",
    "identity information": "identity_information",
    "financial details": "financial_details",
    "health information": "health_information",
    "tax file numbers": "tax_file_numbers",
    "other sensitive information": "other_sensitive_information",
    "consumer data right": "consumer_data_right",
    "digital id information/documents": "digital_id",
}

TIME_BUCKET_MAP: Dict[str, str] = {
    "<=10": "<= 10 days",
    "<= 10": "<= 10 days",
    "11-20": "11-20 days",
    "21-30": "21-30 days",
    ">30": "> 30 days",
    "> 30": "> 30 days",
}

_NUM_RE = re.compile(r"^-?[\d,]+$")


class OAICDataGovError(Exception):
    """Raised when the data.gov.au dataset cannot be discovered or parsed."""


# ----------------------------------------------------------------------
# Small parsing helpers
# ----------------------------------------------------------------------

def _clean(value: Any) -> str:
    """Normalise a cell value to a trimmed string (non-breaking spaces too)."""
    if value is None:
        return ""
    return str(value).replace(" ", " ").strip()


def _as_int(value: Any) -> Optional[int]:
    """Coerce a cell value to int, tolerating comma grouping. None if not numeric."""
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return int(value)
    s = _clean(value)
    if not s or not _NUM_RE.match(s):
        return None
    try:
        return int(s.replace(",", ""))
    except ValueError:
        return None


def _normalize_range_label(label: str) -> str:
    """Canonicalise an individuals-affected bucket label.

    Strips the spaces OAIC puts around the hyphen ("2 - 10" -> "2-10") so the
    label matches the midpoint/bounds tables in build_static_dashboard, fixes
    the "250,000 - 500,000" typo in the Jul-Dec 2025 workbook (every other
    bucket starts at n+1), and maps the open-ended top bucket to a "+" form.
    """
    s = _clean(label)
    if not s:
        return s
    if s.lower() == "unknown":
        return "Unknown"

    m = re.match(r"^([\d,]+)\s*or more$", s, flags=re.IGNORECASE)
    if m:
        return f"{m.group(1)}+"

    s = re.sub(r"\s*-\s*", "-", s)
    if s == "250,000-500,000":  # OAIC typo: should be 250,001
        return "250,001-500,000"
    return s


def _normalize_time_bucket(label: str) -> Optional[str]:
    """Map a workbook time bucket ("≤10", ">30") to the project's canonical form."""
    s = _clean(label).replace("≤", "<=").replace("≥", ">=")
    s = re.sub(r"\s+", "", s)
    return TIME_BUCKET_MAP.get(s)


def _iter_label_count_rows(worksheet) -> List[Tuple[str, Optional[int]]]:
    """Yield (label, count) for every non-empty row of a two-column sheet.

    Consecutive byte-identical rows are collapsed: the Jul-Dec 2025
    "Source of breach" sheet repeats its final row, which would otherwise
    double-count that sub-source.
    """
    rows: List[Tuple[str, Optional[int]]] = []
    for raw in worksheet.iter_rows(values_only=True):
        if not raw:
            continue
        label = _clean(raw[0])
        if not label:
            continue
        count = _as_int(raw[1]) if len(raw) > 1 else None
        if rows and rows[-1] == (label, count):
            continue  # duplicated row in the source workbook
        rows.append((label, count))
    return rows


# ----------------------------------------------------------------------
# CKAN discovery
# ----------------------------------------------------------------------

def discover_resources(dataset_id: str = DATASET_ID, timeout: int = 60) -> List[Dict[str, Any]]:
    """Return the XLSX resources published under the NDB dataset on data.gov.au."""
    try:
        response = requests.get(
            CKAN_PACKAGE_URL,
            params={"id": dataset_id},
            headers={"User-Agent": USER_AGENT},
            timeout=timeout,
        )
        response.raise_for_status()
        payload = response.json()
    except (requests.RequestException, ValueError) as exc:
        raise OAICDataGovError(f"CKAN discovery failed for {dataset_id!r}: {exc}") from exc

    if not payload.get("success"):
        raise OAICDataGovError(f"CKAN returned success=false for {dataset_id!r}")

    resources = payload.get("result", {}).get("resources", []) or []
    xlsx = [r for r in resources if str(r.get("format", "")).upper() == "XLSX"]
    logger.info("Discovered %d XLSX resource(s) of %d total on data.gov.au",
                len(xlsx), len(resources))
    return xlsx


def parse_period_from_name(name: str) -> Optional[Tuple[int, str, int, int]]:
    """Extract (year, period, start_month, end_month) from a resource name.

    Handles the OAIC naming style "NDB Data 1 July 2025 to 31 Dec 2025".
    Returns None when the name does not describe a recognisable half-year.
    """
    text = _clean(name)
    months = "|".join(MONTH_NAMES) + "|" + "|".join(m[:3] for m in MONTH_NAMES)
    pattern = (
        rf"(?P<m1>{months})[a-z]*\s+(?P<y1>\d{{4}})\s+to\s+"
        rf"\d{{1,2}}\s+(?P<m2>{months})[a-z]*\s+(?P<y2>\d{{4}})"
    )
    match = re.search(pattern, text, flags=re.IGNORECASE)
    if not match:
        return None

    def _month_num(token: str) -> Optional[int]:
        token = token.lower()
        if token in MONTH_TO_NUM:
            return MONTH_TO_NUM[token]
        for full, num in MONTH_TO_NUM.items():
            if full.startswith(token):
                return num
        return None

    start_month = _month_num(match.group("m1"))
    end_month = _month_num(match.group("m2"))
    year = int(match.group("y2"))
    if start_month is None or end_month is None:
        return None
    if int(match.group("y1")) != year:
        logger.warning("Resource %r spans two years; using %d", text, year)

    period = "H1" if end_month <= 6 else "H2"
    return year, period, start_month, end_month


def download_resource(url: str, dest: Path, timeout: int = 120) -> Path:
    """Download a resource URL to ``dest`` and return the path."""
    dest.parent.mkdir(parents=True, exist_ok=True)
    logger.info("Downloading %s", url)
    response = requests.get(url, headers={"User-Agent": USER_AGENT}, timeout=timeout)
    response.raise_for_status()
    dest.write_bytes(response.content)
    logger.info("Saved %s (%d bytes)", dest, dest.stat().st_size)
    return dest


# ----------------------------------------------------------------------
# Workbook parsing
# ----------------------------------------------------------------------

def _parse_ndb_by_month(worksheet) -> Tuple[List[Dict[str, Any]], Dict[str, int], Optional[int]]:
    """Parse the "NDB by month" sheet.

    Returns (monthly_notifications, source_totals, grand_total). Month rows
    carry the period count; the rows beneath each month break it down by
    source, which we accumulate into period-wide source totals.
    """
    monthly: List[Dict[str, Any]] = []
    source_totals: Dict[str, int] = {}
    grand_total: Optional[int] = None

    for label, count in _iter_label_count_rows(worksheet):
        if label.lower() == "grand total":
            grand_total = count
            continue
        if label in SOURCE_LABELS:
            if count is not None:
                source_totals[label] = source_totals.get(label, 0) + count
            continue
        if label.lower() in MONTH_TO_NUM:
            monthly.append({"month": label, "count": count})

    return monthly, source_totals, grand_total


def _parse_source_of_breach(worksheet) -> Tuple[Dict[str, int], Dict[str, Dict[str, int]]]:
    """Parse "Source of breach" into category totals and per-category sub-sources."""
    category_totals: Dict[str, int] = {}
    sub_sources: Dict[str, Dict[str, int]] = {}
    current: Optional[str] = None

    for label, count in _iter_label_count_rows(worksheet):
        if label.lower().startswith("note"):
            continue
        if label in SOURCE_CATEGORIES:
            current = label
            if count is not None:
                category_totals[label] = count
            sub_sources.setdefault(label, {})
            continue
        if current and count is not None:
            sub_sources[current][label] = count

    return category_totals, sub_sources


def _parse_individuals_affected(worksheet) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Optional[int]]:
    """Parse "Individuals affected" into the world-wide distribution and the
    large-scale-Australians table (the second block on the same sheet).
    """
    distribution: List[Dict[str, Any]] = []
    large_scale: List[Dict[str, Any]] = []
    grand_total: Optional[int] = None
    in_large_scale = False

    for label, count in _iter_label_count_rows(worksheet):
        low = label.lower()
        if low.startswith("large-scale"):
            in_large_scale = True
            continue
        if low in ("range", "number of individuals world-wide affected by breaches"):
            continue
        if low.startswith("number of individuals affected"):
            continue
        if low == "grand total":
            grand_total = count
            continue
        entry_range = _normalize_range_label(label)
        if not entry_range:
            continue
        if in_large_scale:
            large_scale.append({"range": entry_range, "current_semester": count})
        else:
            distribution.append({"range": entry_range, "count": count})

    return distribution, large_scale, grand_total


def _parse_personal_information(worksheet) -> Dict[str, Optional[int]]:
    """Parse "Personal information" into the project's category-count fields."""
    result: Dict[str, Optional[int]] = {field: None for field in PERSONAL_INFO_FIELD_MAP.values()}
    for label, count in _iter_label_count_rows(worksheet):
        field = PERSONAL_INFO_FIELD_MAP.get(label.lower())
        if field:
            result[field] = count
    return result


def _parse_top_sectors(worksheet) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    """Parse "Top 5 sectors by source" into top_sectors and the sector x source matrix."""
    top_sectors: List[Dict[str, Any]] = []
    matrix: List[Dict[str, Any]] = []
    current: Optional[Dict[str, Any]] = None

    for label, count in _iter_label_count_rows(worksheet):
        low = label.lower()
        if low.startswith("top ") or low.startswith("note"):
            continue
        if label in SOURCE_LABELS:
            if current is None:
                continue
            if label == "Human error":
                current["human_error"] = count
            elif label == "Malicious or criminal attack":
                current["malicious_or_criminal"] = count
            elif label == "System fault":
                current["system_fault"] = count
            continue
        # A sector heading row.
        top_sectors.append({"sector": label, "notifications": count})
        current = {"sector": label, "human_error": None,
                   "malicious_or_criminal": None, "system_fault": None}
        matrix.append(current)

    return top_sectors, matrix


def _parse_time_buckets(worksheet) -> List[Dict[str, Any]]:
    """Parse a "Time to identify/notify by source" sheet into bucket counts."""
    grouped: Dict[str, int] = {}
    for label, count in _iter_label_count_rows(worksheet):
        if label.lower().startswith(("note", "time taken", "source and")):
            continue
        bucket = _normalize_time_bucket(label)
        if bucket and count is not None:
            grouped[bucket] = grouped.get(bucket, 0) + count

    total = sum(grouped.values())
    entries: List[Dict[str, Any]] = []
    for bucket, count in grouped.items():
        entries.append({
            "bucket": bucket,
            "count": count,
            "current_pct": round(count / total * 100, 1) if total else None,
        })
    return entries


def parse_workbook(path: Path) -> Dict[str, Any]:
    """Parse an OAIC NDB workbook into its constituent sections."""
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sections: Dict[str, Any] = {}

        if "NDB by month" in workbook.sheetnames:
            monthly, source_totals, grand_total = _parse_ndb_by_month(workbook["NDB by month"])
            sections["monthly_notifications"] = monthly
            sections["source_totals"] = source_totals
            sections["grand_total"] = grand_total

        if "Source of breach" in workbook.sheetnames:
            category_totals, sub_sources = _parse_source_of_breach(workbook["Source of breach"])
            sections["source_category_totals"] = category_totals
            sections["source_sub_sources"] = sub_sources

        if "Individuals affected" in workbook.sheetnames:
            dist, large_scale, ia_total = _parse_individuals_affected(
                workbook["Individuals affected"]
            )
            sections["individuals_affected_distribution"] = dist
            sections["large_scale_australians"] = large_scale
            sections["individuals_affected_total"] = ia_total

        if "Personal information" in workbook.sheetnames:
            sections["personal_info_types"] = _parse_personal_information(
                workbook["Personal information"]
            )

        if "Top 5 sectors by source" in workbook.sheetnames:
            top_sectors, matrix = _parse_top_sectors(workbook["Top 5 sectors by source"])
            sections["top_sectors"] = top_sectors
            sections["sector_by_source"] = matrix

        if "Time to identify by source" in workbook.sheetnames:
            sections["time_to_identify"] = _parse_time_buckets(
                workbook["Time to identify by source"]
            )
        if "Time to notify by source" in workbook.sheetnames:
            sections["time_to_notify"] = _parse_time_buckets(
                workbook["Time to notify by source"]
            )

        return sections
    finally:
        workbook.close()


# ----------------------------------------------------------------------
# Record construction
# ----------------------------------------------------------------------

def build_record(
    sections: Dict[str, Any],
    year: int,
    period: str,
    start_month: int,
    end_month: int,
    resource_url: str,
) -> Dict[str, Any]:
    """Assemble a statistics record in the project's OAIC schema.

    Field semantics deliberately follow the existing series so the dashboard's
    time-series charts stay continuous:

    ``cyber_incidents_total`` is set to the malicious-or-criminal-attack count,
    matching how every dashboard-scraped period from 2023 H1 onward populates
    it. OAIC's narrower "Cyber incident" sub-source is preserved separately as
    ``cyber_incident_only`` so the finer figure is not lost.

    The per-method breakdown (phishing/ransomware/hacking/brute force/malware/
    compromised credentials) is left as None: data.gov.au does not publish it,
    and inventing values would corrupt those series.
    """
    source_totals: Dict[str, int] = sections.get("source_totals", {}) or {}
    category_totals: Dict[str, int] = sections.get("source_category_totals", {}) or {}
    sub_sources: Dict[str, Dict[str, int]] = sections.get("source_sub_sources", {}) or {}

    total = sections.get("grand_total") or sections.get("individuals_affected_total")

    # Month-level breakdown is the more complete source (it includes the
    # "Other"/"Currently unknown" buckets); fall back to the category sheet.
    malicious = source_totals.get("Malicious or criminal attack") \
        or category_totals.get("Malicious or criminal attack")
    human_error = source_totals.get("Human error") or category_totals.get("Human error")
    system_faults = source_totals.get("System fault") or category_totals.get("System fault")

    cyber_incident_only = (sub_sources.get("Malicious or criminal attack", {}) or {}).get(
        "Cyber incident"
    )

    cyber_pct = None
    if malicious and total:
        cyber_pct = round(malicious / total * 100, 1)

    month_label = f"{MONTH_NAMES[start_month - 1][:3]}-{MONTH_NAMES[end_month - 1][:3]} {year}"

    return {
        "title": f"Notifiable Data Breaches Report: {month_label}",
        "url": DASHBOARD_URL,
        "year": year,
        "period": period,
        "quarter": period,
        "start_month": start_month,
        "end_month": end_month,
        "source": "data.gov.au",
        "resource_url": resource_url,
        "scraped_at": datetime.now().isoformat(),
        "total_notifications": total,
        "malicious_attacks": malicious,
        "human_error": human_error,
        "system_faults": system_faults,
        "other_sources": source_totals.get("Other"),
        "currently_unknown_sources": source_totals.get("Currently unknown"),
        "cyber_incidents_total": malicious,
        "cyber_incidents_percentage": cyber_pct,
        "cyber_incident_only": cyber_incident_only,
        # Not published on data.gov.au - left null rather than guessed.
        "phishing": None,
        "ransomware": None,
        "hacking": None,
        "brute_force": None,
        "malware": None,
        "compromised_credentials": None,
        "monthly_notifications": sections.get("monthly_notifications", []),
        "top_sectors": sections.get("top_sectors", []),
        "sector_by_source": sections.get("sector_by_source", []),
        "individuals_affected_distribution": sections.get(
            "individuals_affected_distribution", []
        ),
        "large_scale_australians": sections.get("large_scale_australians", []),
        "personal_info_types": sections.get("personal_info_types", {}),
        "breach_sub_sources": sub_sources,
        "time_to_identify": sections.get("time_to_identify", []),
        "time_to_notify": sections.get("time_to_notify", []),
        "key_findings": [],
        "pdf_url": None,
        "pdf_parsed": False,
    }


def verify_record(record: Dict[str, Any]) -> List[str]:
    """Return a list of internal-consistency problems found in a record.

    Checks the sums the workbook lets us cross-validate: source counts against
    the period total, and the individuals-affected distribution against it too.
    """
    problems: List[str] = []
    total = record.get("total_notifications")
    if not isinstance(total, int) or total <= 0:
        problems.append("total_notifications missing or non-positive")
        return problems

    parts = [
        record.get("malicious_attacks") or 0,
        record.get("human_error") or 0,
        record.get("system_faults") or 0,
        record.get("other_sources") or 0,
        record.get("currently_unknown_sources") or 0,
    ]
    source_sum = sum(parts)
    if source_sum != total:
        problems.append(f"source breakdown sums to {source_sum}, expected {total}")

    monthly = [m.get("count") or 0 for m in record.get("monthly_notifications", [])]
    if monthly and sum(monthly) != total:
        problems.append(f"monthly counts sum to {sum(monthly)}, expected {total}")

    dist_sum = sum(d.get("count") or 0 for d in record.get("individuals_affected_distribution", []))
    if dist_sum and dist_sum != total:
        problems.append(f"individuals-affected distribution sums to {dist_sum}, expected {total}")

    for entry in record.get("top_sectors", []):
        count = entry.get("notifications")
        if isinstance(count, int) and count > total:
            problems.append(f"sector {entry.get('sector')!r} count {count} exceeds total {total}")

    return problems


# ----------------------------------------------------------------------
# Merge + output
# ----------------------------------------------------------------------

def merge_into_existing(
    new_records: List[Dict[str, Any]],
    existing_path: Optional[Path],
) -> List[Dict[str, Any]]:
    """Merge new records into an existing statistics file, replacing by period."""
    merged: Dict[Tuple[Any, Any], Dict[str, Any]] = {}

    if existing_path:
        try:
            existing = json.loads(existing_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError) as exc:
            raise OAICDataGovError(f"Could not read {existing_path}: {exc}") from exc
        for record in existing:
            merged[(record.get("year"), record.get("period"))] = record
        logger.info("Loaded %d existing period(s) from %s", len(merged), existing_path)

    for record in new_records:
        key = (record.get("year"), record.get("period"))
        if key in merged:
            logger.info("Replacing existing record for %s %s", *key)
        else:
            logger.info("Adding new record for %s %s", *key)
        merged[key] = record

    return sorted(merged.values(), key=lambda r: (r.get("year", 0), r.get("period", "")))


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(
        description="Ingest OAIC NDB statistics published as XLSX on data.gov.au"
    )
    parser.add_argument("--existing-data", type=Path, default=None,
                        help="Existing oaic_cyber_statistics JSON to merge into.")
    parser.add_argument("--output", type=Path, default=None,
                        help="Output filename (default: oaic_cyber_statistics_<timestamp>.json).")
    parser.add_argument("--cache-dir", type=Path, default=Path("instance/oaic_datagov"),
                        help="Where downloaded workbooks are stored.")
    parser.add_argument("--dry-run", action="store_true",
                        help="Parse and report without writing an output file.")
    parser.add_argument("--verbose", action="store_true", help="Enable debug logging.")
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s",
    )

    try:
        resources = discover_resources()
    except OAICDataGovError as exc:
        logger.error("%s", exc)
        return 1

    if not resources:
        logger.warning("No XLSX resources found on data.gov.au; nothing to ingest.")
        return 0

    records: List[Dict[str, Any]] = []
    for resource in resources:
        name = resource.get("name") or ""
        parsed_period = parse_period_from_name(name)
        if not parsed_period:
            logger.warning("Skipping resource with unrecognised period: %r", name)
            continue

        year, period, start_month, end_month = parsed_period
        url = resource.get("url") or ""
        safe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", Path(url).name or f"{year}_{period}.xlsx")
        dest = args.cache_dir / safe_name

        try:
            download_resource(url, dest)
            sections = parse_workbook(dest)
        except (requests.RequestException, OSError, ValueError) as exc:
            logger.error("Failed to ingest %r: %s", name, exc)
            continue

        record = build_record(sections, year, period, start_month, end_month, url)
        problems = verify_record(record)
        if problems:
            for problem in problems:
                logger.warning("[%s %s] %s", year, period, problem)
        else:
            logger.info("[%s %s] all internal consistency checks passed", year, period)

        logger.info(
            "[%s %s] total=%s malicious=%s human_error=%s system_faults=%s sectors=%d",
            year, period, record["total_notifications"], record["malicious_attacks"],
            record["human_error"], record["system_faults"], len(record["top_sectors"]),
        )
        records.append(record)

    if not records:
        logger.error("No records could be built from data.gov.au resources.")
        return 1

    if args.dry_run:
        print(json.dumps(records, indent=2, ensure_ascii=False))
        logger.info("Dry run - no output written.")
        return 0

    merged = merge_into_existing(records, args.existing_data)

    output = args.output or Path(
        f"oaic_cyber_statistics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    )
    output.write_text(json.dumps(merged, indent=2, ensure_ascii=False), encoding="utf-8")
    logger.info("Wrote %d period(s) to %s", len(merged), output)
    return 0


if __name__ == "__main__":
    sys.exit(main())
